#!/usr/bin/env python3
import os
import openpyxl
from pathlib import Path

folder = Path("drive-download-20260618T034045Z-3-001")

for f in sorted(folder.iterdir()):
    if f.suffix.lower() not in (".xlsx", ".xls"):
        continue
    print(f"\n{'='*60}")
    print(f"FILE: {f.name}")
    print('='*60)
    try:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames[:20]}{'...' if len(wb.sheetnames) > 20 else ''}")
        for sheet_name in wb.sheetnames[:5]:
            ws = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                print("  (empty)")
                continue
            # Print first 8 rows
            for i, row in enumerate(rows[:8]):
                cells = [str(c)[:40] if c is not None else "" for c in row[:10]]
                print(f"  R{i+1}: {cells}")
            print(f"  ... total rows: {len(rows)}, dimensions: {ws.dimensions}")
    except Exception as e:
        print(f"  ERROR: {e}")
